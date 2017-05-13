# -*- coding: utf-8 -*-
import re
import copy
from openpyxl.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter


# code from http://stackoverflow.com/questions/17299364/insert-row-into-excel-spreadsheet-using-openpyxl-in-python
def insert_rows(self, row_idx, cnt, above=False, copy_style=True, fill_formulae=True):
    """Inserts new (empty) rows into worksheet at specified row index.

    :param row_idx: Row index specifying where to insert new rows.
    :param cnt: Number of rows to insert.
    :param above: Set True to insert rows above specified row index.
    :param copy_style: Set True if new rows should copy style of immediately above row.
    :param fill_formulae: Set True if new rows should take on formula from immediately above row, filled with references new to rows.

    Usage:

    * insert_rows(2, 10, above=True, copy_style=False)

    """
    CELL_RE  = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

    row_idx = row_idx - 1 if above else row_idx

    def replace(m):
        row = m.group('row')
        prefix = "$" if row.find("$") != -1 else ""
        row = int(row.replace("$",""))
        row += cnt if row > row_idx else 0
        return m.group('col') + prefix + str(row)

    # First, we shift all cells down cnt rows...
    old_cells = set()
    old_fas   = set()
    new_cells = dict()
    new_fas   = dict()
    for c in self._cells.values():

        old_coor = c.coordinate

        # Shift all references to anything below row_idx
        if c.data_type == Cell.TYPE_FORMULA:
            c.value = CELL_RE.sub(
                replace,
                c.value
            )
            # Here, we need to properly update the formula references to reflect new row indices
            if old_coor in self.formula_attributes and 'ref' in self.formula_attributes[old_coor]:
                self.formula_attributes[old_coor]['ref'] = CELL_RE.sub(
                    replace,
                    self.formula_attributes[old_coor]['ref']
                )

        # Do the magic to set up our actual shift    
        if c.row > row_idx:
            old_coor = c.coordinate
            old_cells.add((c.row,c.col_idx))
            c.row += cnt
            new_cells[(c.row,c.col_idx)] = c
            if old_coor in self.formula_attributes:
                old_fas.add(old_coor)
                fa = self.formula_attributes[old_coor].copy()
                new_fas[c.coordinate] = fa

    for coor in old_cells:
        del self._cells[coor]
    self._cells.update(new_cells)

    for fa in old_fas:
        del self.formula_attributes[fa]
    self.formula_attributes.update(new_fas)

    # Next, we need to shift all the Row Dimensions below our new rows down by cnt...
    for row in range(len(self.row_dimensions)-1+cnt,row_idx+cnt,-1):
        new_rd = copy.copy(self.row_dimensions[row-cnt])
        new_rd.index = row
        self.row_dimensions[row] = new_rd
        del self.row_dimensions[row-cnt]

    # Now, create our new rows, with all the pretty cells
    row_idx += 1
    for row in range(row_idx,row_idx+cnt):
        # Create a Row Dimension for our new row
        new_rd = copy.copy(self.row_dimensions[row-1])
        new_rd.index = row
        self.row_dimensions[row] = new_rd
        for col in range(1,self.max_column):
            col = get_column_letter(col)
            cell = self.cell('%s%d'%(col,row))
            cell.value = None
            source = self.cell('%s%d'%(col,row-1))
            if copy_style:
                cell.number_format = source.number_format
                cell.font      = source.font.copy()
                cell.alignment = source.alignment.copy()
                cell.border    = source.border.copy()
                cell.fill      = source.fill.copy()
            if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
                s_coor = source.coordinate
                if s_coor in self.formula_attributes and 'ref' not in self.formula_attributes[s_coor]:
                    fa = self.formula_attributes[s_coor].copy()
                    self.formula_attributes[cell.coordinate] = fa
                # print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
                cell.value = re.sub(
                    "(\$?[A-Z]{1,3}\$?)%d"%(row - 1),
                    lambda m: m.group(1) + str(row),
                    source.value
                )   
                cell.data_type = Cell.TYPE_FORMULA

    # Check for Merged Cell Ranges that need to be expanded to contain new cells
    for cr_idx, cr in enumerate(self.merged_cell_ranges):
        self.merged_cell_ranges[cr_idx] = CELL_RE.sub(
            replace,
            cr
        )

Worksheet.insert_rows = insert_rows
