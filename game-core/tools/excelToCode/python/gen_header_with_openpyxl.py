# -*- coding: utf-8 -*-
import openpyxl
import openpyxl_extend
import xlsconfig
from openpyxl.styles import PatternFill
from tps.tp0 import type2string

header_fill = PatternFill(patternType="solid", fgColor="FABF8F")
type_fill = PatternFill(patternType="solid", fgColor="82C3D4")
field_fill = PatternFill(patternType="solid", fgColor="F0F000")

def create_header(input_file, converter, sheet_index):
	# 索引从1开始
	header_row = xlsconfig.SHEET_ROW_INDEX["header"] + 1
	field_row = xlsconfig.SHEET_ROW_INDEX.get("field", -1) + 1
	type_row = xlsconfig.SHEET_ROW_INDEX.get("type", -1) + 1

	book = openpyxl.load_workbook(input_file)
	table = book.worksheets[sheet_index]

	headers = {}
	for i, cell in enumerate(table[header_row]):
		header = cell.value
		if header is None or header == "": break
		headers[header] = i + 1

	hidden_fileds = getattr(converter, "HIDDEN_FIELDS", ())

	header_2_field = {}
	new_headers = []
	for info in converter.CONFIG:
		header = info[0].decode("utf-8")
		field = info[1]
		type = type2string(info[2])

		header_2_field[header] = field

		if header in headers: continue
		if field in hidden_fileds: continue

		new_headers.append((header, field, type))

	if field_row > 0 and table.cell(row = field_row, column = 1).value != "ID":
		table.insert_rows(field_row, 1, True)
		for header, column in headers.iteritems():
			field = header_2_field.get(header)
			if field is None: continue

			cell = table.cell(row = field_row, column = column, value = field)
			cell.fill = field_fill

	elif len(new_headers) == 0:
		return True

	new_headers.sort(key = lambda x: x[0])
	col = len(headers) + 1
	for header, field, type in new_headers:
		cell = table.cell(row = header_row, column = col, value = header)
		cell.fill = header_fill

		if field_row > 0:
			cell = table.cell(row = field_row, column = col, value = field)
			cell.fill = field_fill

		if type_row > 0:
			cell = table.cell(row = type_row, column = col, value = type)
			cell.fill = type_fill

		col += 1

	book.save(input_file)
	print "生成表头：", input_file
	return True
